# app.py
from datetime import datetime, date, time
from io import BytesIO

import openpyxl
from flask import (
    Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
)
from sqlalchemy import func

from config import Config
from models import (
    db,
    Product,
    ProductAlias,
    StockMovement,
    ImportBatch,
    OnwayItem,
    IncomingItem,
)
from services.excel_import import read_alias_counts_from_excel, file_hash

# --- AI Chatbot (Gemini) ---
import os
from google import genai
from google.genai import types
app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)
print("CURRENT DIR:", os.getcwd())
print("DB:", app.config["SQLALCHEMY_DATABASE_URI"])

def _get_gemini_client() -> genai.Client:
    """Server-side Gemini client. Requires GEMINI_API_KEY (or GOOGLE_API_KEY) env var."""
    # Client will also auto-pick GEMINI_API_KEY/GOOGLE_API_KEY from environment.
    return genai.Client()


CHATBOT_INSTRUCTIONS = os.environ.get(
    "CHATBOT_INSTRUCTIONS",
    """You are a helpful assistant inside the Kazpi Stock MVP web app.
Answer briefly and practically. If the user asks about the app, explain what to click and what it does.
Language: use the same language as the user (Kazakh/Russian).
""",
)

GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-3-flash-preview")


# ----------------------
# Utils
# ----------------------
def normalize_alias(value: str) -> str:
    value = str(value or "").strip().lower()
    return " ".join(value.split())


def calc_delivery_cost_by_price(price: int) -> int:
    p = int(price or 0)
    if p <= 1000:
        return 49
    if p <= 3000:
        return 179
    if p <= 5000:
        return 231
    if p <= 10000:
        return 957
    return 1507


def calc_onway_profit(sale_price: int, purchase_price: int) -> tuple[int, int, int]:
    sale = int(sale_price or 0)
    buy = int(purchase_price or 0)
    commission = int(round(sale * 12 / 100.0))
    delivery = calc_delivery_cost_by_price(sale)
    profit = sale - commission - delivery - buy
    return profit, commission, delivery


def consume_onway_items(product_id: int, qty: int) -> int:
    """Excel SALE кезінде: жолда тұрған позицияларды is_active=False қылып жауып тастайды."""
    if qty <= 0:
        return 0
    items = (
        OnwayItem.query
        .filter_by(product_id=product_id, is_active=True)
        .order_by(OnwayItem.created_at.asc())
        .limit(qty)
        .all()
    )
    for it in items:
        it.is_active = False
    return len(items)


def ensure_sqlite_columns():
    """SQLite-та баған жетіспесе, ALTER TABLE (MVP)."""
    try:
        engine = db.engine
        if engine.dialect.name != "sqlite":
            return

        def has_col(table: str, col: str) -> bool:
            _ = db.session.execute(func.sqlite_version()).all()
            info = db.session.execute(db.text(f"PRAGMA table_info({table})")).fetchall()
            return any(r[1] == col for r in info)

        if not has_col("products", "onway_qty"):
            db.session.execute(db.text("ALTER TABLE products ADD COLUMN onway_qty INTEGER NOT NULL DEFAULT 0"))

        if not has_col("products", "reserved_qty"):
            db.session.execute(db.text("ALTER TABLE products ADD COLUMN reserved_qty INTEGER NOT NULL DEFAULT 0"))

        if not has_col("stock_movements", "sale_sum"):
            db.session.execute(db.text("ALTER TABLE stock_movements ADD COLUMN sale_sum INTEGER NOT NULL DEFAULT 0"))
        if not has_col("stock_movements", "delivery_cost"):
            db.session.execute(db.text("ALTER TABLE stock_movements ADD COLUMN delivery_cost INTEGER NOT NULL DEFAULT 0"))
        if not has_col("stock_movements", "commission_pct"):
            db.session.execute(db.text("ALTER TABLE stock_movements ADD COLUMN commission_pct INTEGER NOT NULL DEFAULT 12"))

        db.session.commit()
    except Exception:
        db.session.rollback()
        return


with app.app_context():
    db.create_all()
    ensure_sqlite_columns()


# ======================
# HOME
# ======================
@app.route("/")
def index():
    products = Product.query.order_by(Product.name.asc()).all()

    # ✅ 1) Складтағы жалпы сумма = (қолжетімді * закуп)
    stock_total = 0
    for p in products:
        stock_qty = int(p.stock_qty or 0)
        onway_qty = int(getattr(p, "onway_qty", 0) or 0)
        available = max(0, stock_qty - onway_qty)
        stock_total += available * int(p.purchase_price or 0)

    # ✅ 2) Келе жатыр (Incoming) жалпы баға
    incoming_total = int(
        db.session.query(func.coalesce(func.sum(IncomingItem.qty * IncomingItem.purchase_price), 0))
        .filter(IncomingItem.is_active == True)  # noqa: E712
        .scalar()
        or 0
    )

    # ✅ 3) Жолда (Onway) жалпы сумма (active sale_price)
    onway_total = int(
        db.session.query(func.coalesce(func.sum(OnwayItem.sale_price), 0))
        .filter(OnwayItem.is_active == True)  # noqa: E712
        .scalar()
        or 0
    )

    # ✅ 4) Жалпы айналым
    turnover_total = int(stock_total) + int(incoming_total) + int(onway_total)

    return render_template(
        "index.html",
        products=products,
        stock_total=stock_total,
        incoming_total=incoming_total,
        onway_total=onway_total,
        turnover_total=turnover_total,
    )


# ======================
# PRODUCTS / STOCK
# ======================
@app.route("/products", methods=["GET", "POST"])
def products():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        purchase_price = request.form.get("purchase_price", "").strip()
        stock_qty = request.form.get("stock_qty", "").strip()

        if not name:
            flash("Тауар атауын енгізіңіз", "error")
            return redirect(url_for("products"))

        if Product.query.filter_by(name=name).first():
            flash("Бұл атаумен тауар бар", "error")
            return redirect(url_for("products"))

        try:
            purchase_price = int(purchase_price)
            stock_qty = int(stock_qty)
        except ValueError:
            flash("Баға мен қалдық бүтін сан болуы керек", "error")
            return redirect(url_for("products"))

        if purchase_price <= 0 or stock_qty < 0:
            flash("Дұрыс мән енгізіңіз", "error")
            return redirect(url_for("products"))

        p = Product(
            name=name,
            purchase_price=purchase_price,
            stock_qty=stock_qty,
            onway_qty=0,
        )
        db.session.add(p)
        db.session.commit()

        flash("Тауар қосылды", "success")
        return redirect(url_for("products"))

    products_list = Product.query.order_by(Product.name.asc()).all()

    # ✅ Складтың жалпы сомасы: "қолжетімді" * закуп
    total_value = 0
    for p in products_list:
        stock_qty = int(p.stock_qty or 0)
        onway_qty = int(getattr(p, "onway_qty", 0) or 0)
        available = max(0, stock_qty - onway_qty)
        total_value += available * int(p.purchase_price or 0)

    return render_template("products.html", products=products_list, total_value=total_value)


@app.route("/products/<int:product_id>", methods=["GET", "POST"])
def product_detail(product_id: int):
    p = Product.query.get_or_404(product_id)

    # ---- POST: alias қосу ----
    if request.method == "POST":
        alias = request.form.get("alias", "").strip()
        if not alias:
            flash("Alias енгізіңіз", "error")
            return redirect(url_for("product_detail", product_id=product_id))

        alias_norm = normalize_alias(alias)

        exists = ProductAlias.query.filter_by(alias=alias_norm).first()
        if exists:
            flash("Бұл alias бұрын қосылған", "error")
            return redirect(url_for("product_detail", product_id=product_id))

        db.session.add(ProductAlias(product_id=p.id, alias=alias_norm))
        db.session.commit()

        flash("Alias қосылды", "success")
        return redirect(url_for("product_detail", product_id=product_id))

    aliases = ProductAlias.query.filter_by(product_id=p.id).all()

    start_str = request.args.get("start", "").strip()
    end_str = request.args.get("end", "").strip()

    q = StockMovement.query.filter_by(product_id=p.id)

    if start_str:
        try:
            sd = datetime.strptime(start_str, "%Y-%m-%d").date()
            q = q.filter(StockMovement.created_at >= datetime.combine(sd, time.min))
        except Exception:
            pass

    if end_str:
        try:
            ed = datetime.strptime(end_str, "%Y-%m-%d").date()
            q = q.filter(StockMovement.created_at <= datetime.combine(ed, time.max))
        except Exception:
            pass

    movements = q.order_by(StockMovement.created_at.desc()).limit(200).all()

    return render_template(
        "product_detail.html",
        product=p,
        aliases=aliases,
        movements=movements,
        start=start_str,
        end=end_str
    )


@app.route("/products/<int:product_id>/edit", methods=["GET", "POST"])
def product_edit(product_id: int):
    p = Product.query.get_or_404(product_id)

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        purchase_price = request.form.get("purchase_price", "").strip()
        stock_qty = request.form.get("stock_qty", "").strip()

        if not name:
            flash("Атау бос болмауы керек", "error")
            return redirect(url_for("product_edit", product_id=product_id))

        try:
            purchase_price = int(purchase_price)
            stock_qty = int(stock_qty)
        except ValueError:
            flash("Қате мән", "error")
            return redirect(url_for("product_edit", product_id=product_id))

        if purchase_price <= 0 or stock_qty < 0:
            flash("Дұрыс мән енгізіңіз", "error")
            return redirect(url_for("product_edit", product_id=product_id))

        exists = Product.query.filter(Product.name == name, Product.id != p.id).first()
        if exists:
            flash("Бұл атаумен басқа тауар бар", "error")
            return redirect(url_for("product_edit", product_id=product_id))

        p.name = name
        p.purchase_price = purchase_price
        p.stock_qty = stock_qty

        db.session.commit()
        flash("Өзгертілді", "success")
        return redirect(url_for("products"))

    return render_template("product_edit.html", product=p)


# ======================
# INCOMING (Келе жатыр)
# ======================
@app.route("/incoming", methods=["GET", "POST"])
def incoming():
    products_list = Product.query.order_by(Product.name.asc()).all()

    if request.method == "POST":
        try:
            product_id = int(request.form.get("product_id", "0"))
            qty = int(request.form.get("qty", "1"))
            purchase_price = int(request.form.get("purchase_price", "0"))
        except Exception:
            flash("Қате мән енгізілді", "error")
            return redirect(url_for("incoming"))

        if product_id <= 0:
            flash("Тауар таңдаңыз", "error")
            return redirect(url_for("incoming"))

        if qty <= 0:
            flash("Саны 1-ден үлкен болсын", "error")
            return redirect(url_for("incoming"))

        if purchase_price <= 0:
            flash("Сатып алынған бағасы 0-ден үлкен болсын", "error")
            return redirect(url_for("incoming"))

        p = Product.query.get_or_404(product_id)

        db.session.add(IncomingItem(
            product_id=p.id,
            qty=qty,
            purchase_price=purchase_price,
            is_active=True,
            created_at=datetime.utcnow()
        ))

        db.session.commit()
        flash("Келе жатқанға қосылды", "success")
        return redirect(url_for("incoming"))

    rows_raw = (
        db.session.query(
            IncomingItem.product_id,
            func.coalesce(func.sum(IncomingItem.qty), 0),
            func.coalesce(func.sum(IncomingItem.qty * IncomingItem.purchase_price), 0)
        )
        .filter(IncomingItem.is_active == True)  # noqa: E712
        .group_by(IncomingItem.product_id)
        .all()
    )

    prod_map = {p.id: p for p in products_list}

    rows = []
    total_incoming_qty = 0
    total_incoming_cost = 0

    for pid, qty_sum, cost_sum in rows_raw:
        p = prod_map.get(pid)
        if not p:
            continue

        qty_sum = int(qty_sum or 0)
        cost_sum = int(cost_sum or 0)

        if qty_sum < 1:
            continue

        rows.append({
            "product_id": pid,
            "name": p.name,
            "qty": qty_sum,
            "cost": cost_sum
        })

        total_incoming_qty += qty_sum
        total_incoming_cost += cost_sum

    rows.sort(key=lambda x: x["name"].lower())

    return render_template(
        "incoming.html",
        products=products_list,
        rows=rows,
        total_incoming_qty=total_incoming_qty,
        total_incoming_cost=total_incoming_cost
    )


# ======================
# STOCK IN (✅ тек Incoming-тен)
# ======================
@app.route("/stock-in", methods=["GET", "POST"])
def stock_in():
    incoming_rows = (
        db.session.query(
            IncomingItem.product_id,
            func.coalesce(func.sum(IncomingItem.qty), 0).label("qty_sum")
        )
        .filter(IncomingItem.is_active == True)  # noqa: E712
        .group_by(IncomingItem.product_id)
        .having(func.coalesce(func.sum(IncomingItem.qty), 0) > 0)
        .all()
    )

    incoming_map = {pid: int(qty_sum or 0) for pid, qty_sum in incoming_rows}
    incoming_product_ids = list(incoming_map.keys())

    products_list = []
    if incoming_product_ids:
        products_list = (
            Product.query
            .filter(Product.id.in_(incoming_product_ids))
            .order_by(Product.name.asc())
            .all()
        )

    if request.method == "POST":
        try:
            product_id = int(request.form.get("product_id", "0"))
            qty = int(request.form.get("qty", "0"))
        except Exception:
            flash("Қате мән", "error")
            return redirect(url_for("stock_in"))

        note = (request.form.get("note") or "").strip()

        if product_id <= 0:
            flash("Тауар таңдаңыз", "error")
            return redirect(url_for("stock_in"))

        if qty <= 0:
            flash("Саны 1-ден үлкен болсын", "error")
            return redirect(url_for("stock_in"))

        available_incoming = int(incoming_map.get(product_id, 0))
        if available_incoming <= 0:
            flash("Бұл тауар 'Келе жатыр' тізімінде жоқ", "error")
            return redirect(url_for("stock_in"))

        if qty > available_incoming:
            flash(f"Саны 'Келе жатыр' санынан көп болмауы керек (бар: {available_incoming})", "error")
            return redirect(url_for("stock_in"))

        p = Product.query.get_or_404(product_id)

        # FIFO: incoming-тан азайтамыз
        remaining = qty
        items = (
            IncomingItem.query
            .filter_by(product_id=product_id, is_active=True)
            .order_by(IncomingItem.created_at.asc())
            .all()
        )

        for it in items:
            if remaining <= 0:
                break

            it_qty = int(it.qty or 0)
            if it_qty <= 0:
                it.is_active = False
                continue

            take = min(it_qty, remaining)
            it.qty = it_qty - take
            remaining -= take

            if it.qty <= 0:
                it.is_active = False

        if remaining != 0:
            db.session.rollback()
            flash("Incoming азайту кезінде қате шықты (qty сәйкес емес)", "error")
            return redirect(url_for("stock_in"))

        # складқа қосу
        p.stock_qty = int(p.stock_qty or 0) + qty

        db.session.add(StockMovement(
            product_id=p.id,
            kind="IN",
            qty_change=qty,
            note=note or "Келе жатқаннан келіп түсті",
            created_at=datetime.utcnow()
        ))

        db.session.commit()
        flash("Келіп түсті (Incoming-тан алынды)", "success")
        return redirect(url_for("stock_in"))

    return render_template("stock_in.html", products=products_list, incoming_map=incoming_map)


# ======================
# ONWAY LIST
# ======================
@app.route("/onway")
def onway():
    products_list = (
        Product.query
        .filter(Product.onway_qty >= 1)
        .order_by(Product.name.asc())
        .all()
    )

    purchase_by_pid = {p.id: int(p.purchase_price or 0) for p in products_list}

    sums_by_pid = dict(
        db.session.query(
            OnwayItem.product_id,
            func.coalesce(func.sum(OnwayItem.sale_price), 0)
        )
        .filter(OnwayItem.is_active == True)  # noqa: E712
        .group_by(OnwayItem.product_id)
        .all()
    )

    profit_by_pid = {}
    total_onway_profit = 0

    active_items = (
        db.session.query(OnwayItem.product_id, OnwayItem.sale_price)
        .filter(OnwayItem.is_active == True)  # noqa: E712
        .all()
    )

    for pid, sale_price in active_items:
        if pid not in purchase_by_pid:
            continue

        sale_price_i = int(sale_price or 0)
        purchase_price_i = int(purchase_by_pid.get(pid, 0))
        profit, _commission, _delivery = calc_onway_profit(sale_price_i, purchase_price_i)

        profit_by_pid[pid] = profit_by_pid.get(pid, 0) + profit
        total_onway_profit += profit

    total_onway_qty = sum(int(p.onway_qty or 0) for p in products_list)
    total_onway_sum = sum(int(sums_by_pid.get(p.id, 0) or 0) for p in products_list)

    return render_template(
        "onway.html",
        products=products_list,
        total_onway_qty=total_onway_qty,
        total_onway_sum=total_onway_sum,
        total_onway_profit=total_onway_profit,
        sums_by_pid=sums_by_pid,
        profit_by_pid=profit_by_pid
    )


# ✅ ONWAY ADD (alias dropdown + search)
@app.route("/onway-add", methods=["GET", "POST"])
def onway_add():
    aliases = (
        db.session.query(ProductAlias)
        .order_by(ProductAlias.alias.asc())
        .all()
    )

    if request.method == "POST":
        alias_id_raw = (request.form.get("alias_id") or "").strip()
        sale_price_raw = (request.form.get("sale_price") or "").strip()

        try:
            alias_id = int(alias_id_raw)
        except Exception:
            alias_id = 0

        if alias_id <= 0:
            flash("Карточка атауын таңдаңыз", "error")
            return redirect(url_for("onway_add"))

        if not sale_price_raw:
            flash("Сатылған бағаны енгізіңіз", "error")
            return redirect(url_for("onway_add"))

        try:
            sale_price = int(float(sale_price_raw))
        except Exception:
            flash("Сатылған баға сан болуы керек", "error")
            return redirect(url_for("onway_add"))

        if sale_price <= 0:
            flash("Сатылған баға 0-ден үлкен болу керек", "error")
            return redirect(url_for("onway_add"))

        pa = ProductAlias.query.get_or_404(alias_id)
        p = Product.query.get_or_404(pa.product_id)

        p.onway_qty = int(p.onway_qty or 0) + 1

        db.session.add(OnwayItem(
            product_id=p.id,
            sale_price=sale_price,
            is_active=True,
            created_at=datetime.utcnow()
        ))

        db.session.add(StockMovement(
            product_id=p.id,
            kind="ONWAY",
            qty_change=1,
            sale_sum=sale_price,
            delivery_cost=0,
            commission_pct=0,
            note=f"Қолмен жолдаға қосылды (alias: {pa.alias})",
            created_at=datetime.utcnow()
        ))

        db.session.commit()
        flash("Жолдаға қосылды (+1)", "success")
        return redirect(url_for("onway"))

    return render_template("onway_add.html", aliases=aliases)


# ✅ ONWAY LIST -1 (өнім бойынша соңғы позицияны алып тастау)
@app.route("/onway-remove/<int:product_id>", methods=["POST"])
def onway_remove(product_id: int):
    p = Product.query.get_or_404(product_id)

    if int(p.onway_qty or 0) <= 0:
        flash("Бұл тауар жолдада жоқ", "error")
        return redirect(url_for("onway"))

    last_item = (
        OnwayItem.query
        .filter_by(product_id=p.id, is_active=True)
        .order_by(OnwayItem.created_at.desc())
        .first()
    )
    if not last_item:
        p.onway_qty = 0
        db.session.commit()
        flash("Жолда деректері сәйкес емес еді, 0-ге түсті", "error")
        return redirect(url_for("onway"))

    last_item.is_active = False
    p.onway_qty = max(0, int(p.onway_qty or 0) - 1)

    db.session.add(StockMovement(
        product_id=p.id,
        kind="ONWAY",
        qty_change=-1,
        sale_sum=int(last_item.sale_price or 0),
        delivery_cost=0,
        commission_pct=0,
        note="Қолмен жолдадан алынды (-1)",
        created_at=datetime.utcnow()
    ))

    db.session.commit()
    flash("Жолдадан алынды (-1)", "success")
    return redirect(url_for("onway"))


# ✅ ONWAY PRODUCT DETAIL
@app.route("/onway/<int:product_id>")
def onway_product_detail(product_id: int):
    p = Product.query.get_or_404(product_id)

    items = (
        OnwayItem.query
        .filter_by(product_id=p.id, is_active=True)
        .order_by(OnwayItem.created_at.desc())
        .all()
    )

    rows = []
    total_sum = 0
    total_profit = 0
    total_commission = 0
    total_delivery = 0

    for x in items:
        sale_price = int(x.sale_price or 0)
        profit, commission, delivery = calc_onway_profit(sale_price, int(p.purchase_price or 0))

        rows.append({
            "id": x.id,
            "created_at": x.created_at,
            "sale_price": sale_price,
            "commission": commission,
            "delivery": delivery,
            "purchase_price": int(p.purchase_price or 0),
            "profit": profit
        })

        total_sum += sale_price
        total_profit += profit
        total_commission += commission
        total_delivery += delivery

    return render_template(
        "onway_product_detail.html",
        product=p,
        rows=rows,
        total_sum=total_sum,
        total_profit=total_profit,
        total_commission=total_commission,
        total_delivery=total_delivery
    )


# ✅ ONWAY ITEM CANCEL (нақты позицияны -1)
@app.route("/onway/item/<int:item_id>/cancel", methods=["POST"])
def onway_item_cancel(item_id: int):
    item = OnwayItem.query.get_or_404(item_id)

    if not item.is_active:
        flash("Бұл позиция бұрын алып тасталған", "error")
        return redirect(url_for("onway_product_detail", product_id=item.product_id))

    p = Product.query.get_or_404(item.product_id)

    item.is_active = False
    p.onway_qty = max(0, int(p.onway_qty or 0) - 1)

    db.session.add(StockMovement(
        product_id=p.id,
        kind="ONWAY",
        qty_change=-1,
        sale_sum=int(item.sale_price or 0),
        delivery_cost=0,
        commission_pct=0,
        note="Onway позициясы отмена (-1)",
        created_at=datetime.utcnow()
    ))

    db.session.commit()
    flash("Отмена жасалды (-1)", "success")
    return redirect(url_for("onway_product_detail", product_id=p.id))


# ======================
# EXCEL SALES IMPORT
# ======================
def parse_kaspi_archive_orders(file_bytes: bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, 120)]
    header_map = {str(h).strip(): i for i, h in enumerate(headers, start=1) if h}

    required = [
        "Название товара в Kaspi Магазине",
        "Статус",
        "Количество",
        "Сумма",
        "Стоимость доставки для продавца"
    ]
    if not all(k in header_map for k in required):
        return None

    col_name = header_map["Название товара в Kaspi Магазине"]
    col_status = header_map["Статус"]
    col_qty = header_map["Количество"]
    col_sum = header_map["Сумма"]
    col_delivery = header_map["Стоимость доставки для продавца"]

    result = {}
    for r in range(2, ws.max_row + 1):
        status = ws.cell(r, col_status).value
        if not status:
            continue
        status = normalize_alias(status)

        if status not in ["выдан", "доставлен", "завершен", "завершён"]:
            continue

        name = ws.cell(r, col_name).value
        if not name:
            continue
        alias = normalize_alias(name)

        qty_val = ws.cell(r, col_qty).value or 0
        sum_val = ws.cell(r, col_sum).value or 0
        del_val = ws.cell(r, col_delivery).value or 0

        try:
            qty_i = int(float(qty_val))
        except Exception:
            qty_i = 0

        try:
            sum_i = int(round(float(sum_val)))
        except Exception:
            sum_i = 0

        try:
            del_i = int(round(float(del_val)))
        except Exception:
            del_i = 0

        if qty_i <= 0:
            continue

        if alias not in result:
            result[alias] = {"qty": 0, "sum": 0, "delivery": 0}
        result[alias]["qty"] += qty_i
        result[alias]["sum"] += sum_i
        result[alias]["delivery"] += del_i

    return result


@app.route("/import", methods=["GET", "POST"])
def import_page():
    if request.method == "POST":
        f = request.files.get("file")
        sale_date_str = request.form.get("sale_date", "").strip()

        if not f:
            flash("Файл таңдаңыз", "error")
            return redirect(url_for("import_page"))

        if not sale_date_str:
            flash("Сатылған күнді таңдаңыз", "error")
            return redirect(url_for("import_page"))

        try:
            d = datetime.strptime(sale_date_str, "%Y-%m-%d").date()
        except ValueError:
            flash("Күн форматы қате", "error")
            return redirect(url_for("import_page"))

        fixed_dt = datetime.combine(d, time(12, 0, 0))

        file_bytes = f.read()
        h = file_hash(file_bytes)

        if ImportBatch.query.filter_by(file_hash=h).first():
            flash("Бұл файл бұрын жүктелген", "error")
            return redirect(url_for("import_page"))

        kaspi_map = parse_kaspi_archive_orders(file_bytes)

        updates = {}
        unknown = []

        if kaspi_map is not None:
            for alias, agg in kaspi_map.items():
                pa = ProductAlias.query.filter_by(alias=alias).first()
                if not pa:
                    unknown.append((alias, agg["qty"]))
                    continue
                if pa.product_id not in updates:
                    updates[pa.product_id] = {"qty": 0, "sum": 0, "delivery": 0}
                updates[pa.product_id]["qty"] += agg["qty"]
                updates[pa.product_id]["sum"] += agg["sum"]
                updates[pa.product_id]["delivery"] += agg["delivery"]
        else:
            counts = read_alias_counts_from_excel(file_bytes)
            for alias, qty in counts.items():
                pa = ProductAlias.query.filter_by(alias=alias).first()
                if not pa:
                    unknown.append((alias, qty))
                    continue
                if pa.product_id not in updates:
                    updates[pa.product_id] = {"qty": 0, "sum": 0, "delivery": 0}
                updates[pa.product_id]["qty"] += qty

        if unknown:
            return render_template("import_result.html", unknown=unknown, applied=[], blocked=True)

        # Қорғаныс: (склад + жолда) жетуі керек
        shortage = []
        for pid, agg in updates.items():
            p = Product.query.get(pid)
            qty = int(agg["qty"])
            can_supply = int(p.stock_qty or 0) + int(p.onway_qty or 0)
            if can_supply < qty:
                shortage.append((p.name, int(p.stock_qty or 0), int(p.onway_qty or 0), qty))

        if shortage:
            unknown_like = [
                (f"{name} (склад: {stock}, жолда: {onway})", need)
                for (name, stock, onway, need) in shortage
            ]
            return render_template("import_result.html", unknown=unknown_like, applied=[], blocked=True)

        # APPLY: алдымен жолдадан, қалғаны складтан
        for pid, agg in updates.items():
            p = Product.query.get(pid)
            qty = int(agg["qty"])
            sale_sum = int(agg["sum"])
            delivery = int(agg["delivery"])

            take_from_onway = min(int(p.onway_qty or 0), qty)
            if take_from_onway > 0:
                _ = consume_onway_items(p.id, take_from_onway)
                p.onway_qty = int(p.onway_qty or 0) - take_from_onway

            remaining = qty - take_from_onway
            if remaining > 0:
                p.stock_qty = int(p.stock_qty or 0) - remaining

            db.session.add(StockMovement(
                product_id=p.id,
                kind="SALE",
                qty_change=-qty,
                sale_sum=sale_sum,
                delivery_cost=delivery,
                commission_pct=12,
                note=f"Excel импорт: {f.filename} (onway:{take_from_onway}, stock:{remaining})",
                created_at=fixed_dt
            ))

        db.session.add(ImportBatch(file_name=f.filename, file_hash=h))
        db.session.commit()

        applied = [(pid, agg["qty"]) for pid, agg in updates.items()]
        return render_template("import_result.html", unknown=[], applied=applied, blocked=False)

    return render_template("import.html")


# ======================
# REPORT
# ======================
def _calc_profit_rows(start_d: date, end_d: date):
    start_dt = datetime.combine(start_d, time.min)
    end_dt = datetime.combine(end_d, time.max)

    sales = (
        db.session.query(StockMovement, Product)
        .join(Product, Product.id == StockMovement.product_id)
        .filter(StockMovement.kind == "SALE")
        .filter(StockMovement.created_at >= start_dt)
        .filter(StockMovement.created_at <= end_dt)
        .all()
    )

    by_product = {}
    by_day = {}

    for m, p in sales:
        qty = int(-m.qty_change)
        revenue = int(m.sale_sum or 0)
        delivery = int(m.delivery_cost or 0)
        pct = int(m.commission_pct or 12)

        commission = int(round(revenue * pct / 100.0))
        cogs = qty * int(p.purchase_price or 0)
        profit = revenue - commission - delivery - cogs

        if p.id not in by_product:
            by_product[p.id] = {
                "product_id": p.id,
                "name": p.name,
                "qty": 0,
                "revenue": 0,
                "commission": 0,
                "delivery": 0,
                "cogs": 0,
                "profit": 0
            }
        bp = by_product[p.id]
        bp["qty"] += qty
        bp["revenue"] += revenue
        bp["commission"] += commission
        bp["delivery"] += delivery
        bp["cogs"] += cogs
        bp["profit"] += profit

        day_key = m.created_at.date().isoformat()
        if day_key not in by_day:
            by_day[day_key] = {"day": day_key, "qty": 0, "profit": 0, "revenue": 0}
        by_day[day_key]["qty"] += qty
        by_day[day_key]["profit"] += profit
        by_day[day_key]["revenue"] += revenue

    prod_rows = sorted(by_product.values(), key=lambda x: x["name"].lower())
    day_rows = sorted(by_day.values(), key=lambda x: x["day"])

    totals = {
        "qty": sum(r["qty"] for r in prod_rows),
        "revenue": sum(r["revenue"] for r in prod_rows),
        "commission": sum(r["commission"] for r in prod_rows),
        "delivery": sum(r["delivery"] for r in prod_rows),
        "cogs": sum(r["cogs"] for r in prod_rows),
        "profit": sum(r["profit"] for r in prod_rows),
    }

    return prod_rows, day_rows, totals


@app.route("/report")
def report():
    today = datetime.utcnow().date()
    start_str = request.args.get("start", "")
    end_str = request.args.get("end", "")

    if start_str and end_str:
        try:
            start_d = datetime.strptime(start_str, "%Y-%m-%d").date()
            end_d = datetime.strptime(end_str, "%Y-%m-%d").date()
        except ValueError:
            start_d = today.replace(day=1)
            end_d = today
    else:
        start_d = today.replace(day=1)
        end_d = today

    prod_rows, day_rows, totals = _calc_profit_rows(start_d, end_d)

    return render_template(
        "report.html",
        start=start_d.isoformat(),
        end=end_d.isoformat(),
        prod_rows=prod_rows,
        day_rows=day_rows,
        totals=totals
    )


@app.route("/report/export")
def report_export():
    today = datetime.utcnow().date()
    start_str = request.args.get("start", "")
    end_str = request.args.get("end", "")

    try:
        start_d = datetime.strptime(start_str, "%Y-%m-%d").date()
        end_d = datetime.strptime(end_str, "%Y-%m-%d").date()
    except Exception:
        start_d = today.replace(day=1)
        end_d = today

    prod_rows, day_rows, totals = _calc_profit_rows(start_d, end_d)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "By Product"

    ws1.append(["Тауар", "Сатылды (дана)", "Түсім", "Комиссия 12%", "Доставка (продавец)", "Закуп (COGS)", "Таза пайда"])
    for r in prod_rows:
        ws1.append([r["name"], r["qty"], r["revenue"], r["commission"], r["delivery"], r["cogs"], r["profit"]])

    ws1.append([])
    ws1.append(["ИТОГО", totals["qty"], totals["revenue"], totals["commission"], totals["delivery"], totals["cogs"], totals["profit"]])

    ws2 = wb.create_sheet("By Day")
    ws2.append(["Күн", "Сатылды (дана)", "Түсім", "Таза пайда"])
    for r in day_rows:
        ws2.append([r["day"], r["qty"], r["revenue"], r["profit"]])

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"report_{start_d.isoformat()}_{end_d.isoformat()}.xlsx"
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ======================
# API (Android үшін)
# ======================

def _product_to_dict(p: Product) -> dict:
    stock_qty = int(p.stock_qty or 0)
    onway_qty = int(getattr(p, "onway_qty", 0) or 0)
    available_qty = max(0, stock_qty - onway_qty)

    return {
        "id": p.id,
        "name": p.name,
        "purchase_price": int(p.purchase_price or 0),
        "stock_qty": stock_qty,
        "onway_qty": onway_qty,
        "available_qty": available_qty,
    }


@app.get("/api/summary")
def api_summary():
    # 1) Складтағы жалпы сумма = (қолжетімді * закуп)
    products = Product.query.all()
    stock_total = 0
    for p in products:
        stock_qty = int(p.stock_qty or 0)
        onway_qty = int(getattr(p, "onway_qty", 0) or 0)
        available = max(0, stock_qty - onway_qty)
        stock_total += available * int(p.purchase_price or 0)

    # 2) Келе жатыр (Incoming) жалпы баға
    incoming_total = int(
        db.session.query(func.coalesce(func.sum(IncomingItem.qty * IncomingItem.purchase_price), 0))
        .filter(IncomingItem.is_active == True)  # noqa: E712
        .scalar()
        or 0
    )

    # 3) Жолда (Onway) жалпы сумма (active sale_price)
    onway_total = int(
        db.session.query(func.coalesce(func.sum(OnwayItem.sale_price), 0))
        .filter(OnwayItem.is_active == True)  # noqa: E712
        .scalar()
        or 0
    )

    turnover_total = int(stock_total) + int(incoming_total) + int(onway_total)

    return jsonify({
        "stockTotal": stock_total,
        "incomingTotal": incoming_total,
        "onwayTotal": onway_total,
        "turnoverTotal": turnover_total,
    })


# ----------------------
# PRODUCTS
# ----------------------
# ✅ GET /api/products?q=...  (іздеу name бойынша)
@app.get("/api/products")
def api_products():
    q = (request.args.get("q") or "").strip()

    qry = Product.query
    if q:
        # case-insensitive LIKE
        like = f"%{q.lower()}%"
        qry = qry.filter(func.lower(Product.name).like(like))

    rows = qry.order_by(Product.id.desc()).all()
    return jsonify([_product_to_dict(p) for p in rows])


# ✅ POST /api/products  (тауар қосу)
@app.post("/api/products")
def api_products_create():
    data = request.get_json(silent=True) or {}

    name = str(data.get("name") or "").strip()
    purchase_price = data.get("purchase_price", 0)
    stock_qty = data.get("stock_qty", 0)

    if not name:
        return jsonify({"ok": False, "error": "name міндетті"}), 400

    if Product.query.filter_by(name=name).first():
        return jsonify({"ok": False, "error": "Бұл атаумен тауар бар"}), 400

    try:
        purchase_price = int(purchase_price)
        stock_qty = int(stock_qty)
    except Exception:
        return jsonify({"ok": False, "error": "purchase_price және stock_qty сан болу керек"}), 400

    if purchase_price <= 0 or stock_qty < 0:
        return jsonify({"ok": False, "error": "Дұрыс мән енгізіңіз"}), 400

    p = Product(
        name=name,
        purchase_price=purchase_price,
        stock_qty=stock_qty,
        onway_qty=int(getattr(Product, "onway_qty", 0) or 0) if hasattr(Product, "onway_qty") else 0,
    )
    # егер модельде onway_qty бар болса, үстінде anyway 0 береміз:
    if hasattr(p, "onway_qty"):
        p.onway_qty = 0

    db.session.add(p)
    db.session.commit()

    return jsonify({"ok": True, "product": _product_to_dict(p)}), 201


# ✅ PUT /api/products/<id>  (тауар өңдеу)
@app.put("/api/products/<int:product_id>")
def api_products_update(product_id: int):
    p = Product.query.get_or_404(product_id)
    data = request.get_json(silent=True) or {}

    name = str(data.get("name") or p.name).strip()
    purchase_price = data.get("purchase_price", p.purchase_price)
    stock_qty = data.get("stock_qty", p.stock_qty)

    if not name:
        return jsonify({"ok": False, "error": "name бос болмауы керек"}), 400

    exists = Product.query.filter(Product.name == name, Product.id != p.id).first()
    if exists:
        return jsonify({"ok": False, "error": "Бұл атаумен басқа тауар бар"}), 400

    try:
        purchase_price = int(purchase_price)
        stock_qty = int(stock_qty)
    except Exception:
        return jsonify({"ok": False, "error": "purchase_price және stock_qty сан болу керек"}), 400

    if purchase_price <= 0 or stock_qty < 0:
        return jsonify({"ok": False, "error": "Дұрыс мән енгізіңіз"}), 400

    p.name = name
    p.purchase_price = purchase_price
    p.stock_qty = stock_qty
    db.session.commit()

    return jsonify({"ok": True, "product": _product_to_dict(p)})


# ✅ DELETE /api/products/<id>  (тауар өшіру) — керек болса
@app.delete("/api/products/<int:product_id>")
def api_products_delete(product_id: int):
    p = Product.query.get_or_404(product_id)

    # Қауіпсіздік: байланысқан деректер бар ма тексеруге болады
    # Қазір MVP: тікелей өшіреміз
    db.session.delete(p)
    db.session.commit()

    return jsonify({"ok": True})


# ----------------------
# INCOMING (Келе жатыр)
# ----------------------
# ✅ GET /api/incoming : актив incoming тізімі (group by product)
@app.get("/api/incoming")
def api_incoming():
    rows_raw = (
        db.session.query(
            IncomingItem.product_id,
            func.coalesce(func.sum(IncomingItem.qty), 0).label("qty_sum"),
            func.coalesce(func.sum(IncomingItem.qty * IncomingItem.purchase_price), 0).label("cost_sum"),
        )
        .filter(IncomingItem.is_active == True)  # noqa: E712
        .group_by(IncomingItem.product_id)
        .all()
    )

    prod_ids = [pid for pid, _q, _c in rows_raw]
    prod_map = {}
    if prod_ids:
        for p in Product.query.filter(Product.id.in_(prod_ids)).all():
            prod_map[p.id] = p

    out_rows = []
    total_qty = 0
    total_cost = 0

    for pid, qty_sum, cost_sum in rows_raw:
        p = prod_map.get(pid)
        if not p:
            continue
        qty_sum = int(qty_sum or 0)
        cost_sum = int(cost_sum or 0)
        if qty_sum <= 0:
            continue

        out_rows.append({
            "product_id": pid,
            "name": p.name,
            "qty": qty_sum,
            "cost": cost_sum,
        })
        total_qty += qty_sum
        total_cost += cost_sum

    out_rows.sort(key=lambda x: x["name"].lower())

    return jsonify({
        "rows": out_rows,
        "total_qty": total_qty,
        "total_cost": total_cost,
    })


# ----------------------
# ONWAY (Жолда)
# ----------------------
# ✅ GET /api/onway : onway_qty>=1 тізімі (өнім бойынша)
@app.get("/api/onway")
def api_onway():
    products_list = (
        Product.query
        .filter(Product.onway_qty >= 1)
        .order_by(Product.name.asc())
        .all()
    )

    sums_by_pid = dict(
        db.session.query(
            OnwayItem.product_id,
            func.coalesce(func.sum(OnwayItem.sale_price), 0)
        )
        .filter(OnwayItem.is_active == True)  # noqa: E712
        .group_by(OnwayItem.product_id)
        .all()
    )

    out = []
    total_onway_qty = 0
    total_onway_sum = 0

    for p in products_list:
        pid = p.id
        onway_qty = int(p.onway_qty or 0)
        ssum = int(sums_by_pid.get(pid, 0) or 0)

        out.append({
            "product_id": pid,
            "name": p.name,
            "onway_qty": onway_qty,
            "sum": ssum,
        })

        total_onway_qty += onway_qty
        total_onway_sum += ssum

    return jsonify({
        "rows": out,
        "total_onway_qty": total_onway_qty,
        "total_onway_sum": total_onway_sum,
    })


# ======================
# AI CHATBOT
# ======================
@app.route("/chat")
def chat_page():
    """Chat UI page."""
    return render_template("chat.html")


@app.route("/api/chat", methods=["POST"])
def api_chat():
    """JSON API for chat messages.

    Request:
      {"message": "..."}
    Response:
      {"reply": "..."}
    """
    data = request.get_json(silent=True) or {}
    msg = (data.get("message") or "").strip()
    if not msg:
        return jsonify({"error": "Empty message"}), 400

    if not (os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")):
        return jsonify({
            "error": "Server is not configured: set GEMINI_API_KEY env var"
        }), 500

    try:
        client = _get_gemini_client()
        resp = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=msg,
            config=types.GenerateContentConfig(
                system_instruction=CHATBOT_INSTRUCTIONS,
                max_output_tokens=1200,  # 🔥 ұлғайттық (800–2000 қоя аласың)
                temperature=0.6,
            ),
        )

        # 1) resp.text
        text = (getattr(resp, "text", None) or "").strip()

        # 2) толық жинау: candidates -> content.parts
        parts = []
        if getattr(resp, "candidates", None):
            for cand in resp.candidates:
                content = getattr(cand, "content", None)
                if content and getattr(content, "parts", None):
                    for p in content.parts:
                        t = getattr(p, "text", None)
                        if t:
                            parts.append(t)

        joined = "".join(parts).strip()
        if len(joined) > len(text):
            text = joined

        # Диагностика: неге тоқтады?
        finish_reason = ""
        if getattr(resp, "candidates", None):
            finish_reason = str(getattr(resp.candidates[0], "finish_reason", ""))  # STOP / MAX_TOKENS / SAFETY ...
            print("Gemini finish_reason:", finish_reason)

        reply = text or "(empty reply)"
        return jsonify({"reply": reply, "finish_reason": finish_reason})
    except Exception as e:
        return jsonify({"error": f"AI request failed: {e}"}), 500


if __name__ == "__main__":
    # ✅ Эмулятор/телефон кіре алуы үшін:
    # - localhost емес, барлық интерфейсті тыңдаймыз
    app.run(host="0.0.0.0", port=5000, debug=True)
