import hashlib
import os
import tempfile
from collections import Counter
from openpyxl import load_workbook


def normalize(text: str) -> str:
    if not text:
        return ""
    text = str(text).strip().lower()
    text = " ".join(text.split())
    return text


def file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def read_alias_counts_from_excel(file_bytes: bytes, col=1, start_row=2):
    """
    Windows-та NamedTemporaryFile(delete=True) + openpyxl комбинациясы PermissionError беруі мүмкін.
    Сондықтан:
      - delete=False -> файлды жауып тастаймыз
      - openpyxl оқиды
      - соңында өзіміз өшіреміз
    """
    tmp_path = None
    try:
        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        # fd арқылы ашылған дескрипторды жауып тастау маңызды (Windows)
        os.close(fd)

        with open(tmp_path, "wb") as f:
            f.write(file_bytes)

        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active

        values = []
        for r in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=col).value
            alias = normalize(cell)
            if not alias:
                continue
            if alias in ("итого", "барлығы", "всего"):
                continue
            values.append(alias)

        return Counter(values)

    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                # кейде антивирус/индексация ұстап қалуы мүмкін - бұл тестте критикалық емес
                pass
