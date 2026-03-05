from io import BytesIO
import openpyxl

from services.normalize import normalize_alias


def read_onway_from_active_orders_xlsx(file_bytes: bytes):
    """
    ActiveOrders.xlsx файлынан тек "Передан курьеру" статусын оқиды.
    Атауды ТЕК "Название товара в Kaspi Магазине" (3-баған) арқылы алады.
    Қайтарады: onway_map(alias -> qty), error_message
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Header (1-жол)
    headers = []
    for c in range(1, 200):
        v = ws.cell(1, c).value
        headers.append(str(v).strip() if v is not None else "")

    def find_col_exact(text: str):
        t = text.strip().lower()
        for i, h in enumerate(headers, start=1):
            if (h or "").strip().lower() == t:
                return i
        return None

    def find_col_contains(substr: str):
        s = substr.strip().lower()
        for i, h in enumerate(headers, start=1):
            if s in (h or "").lower():
                return i
        return None

    # Міндетті бағандар
    col_status = find_col_contains("статус")
    col_qty = find_col_contains("колич")

    # ✅ Нақты Kaspi бағаны
    col_name = find_col_exact("Название товара в Kaspi Магазине")
    if col_name is None:
        col_name = find_col_contains("в kaspi магазине")
    if col_name is None:
        col_name = find_col_contains("название")

    if not col_status or not col_qty or not col_name:
        return {}, "Файл форматында 'Название товара в Kaspi Магазине/Статус/Количество' бағандары табылмады"

    onway_map = {}

    for r in range(2, ws.max_row + 1):
        status = ws.cell(r, col_status).value
        if not status:
            continue
        status_l = normalize_alias(status)

        # тек "Передан курьеру"
        if status_l != "передан курьеру":
            continue

        name = ws.cell(r, col_name).value
        if not name:
            continue
        alias = normalize_alias(name)

        qty_val = ws.cell(r, col_qty).value or 0
        try:
            qty_i = int(float(qty_val))
        except Exception:
            qty_i = 0

        if qty_i <= 0:
            continue

        onway_map[alias] = onway_map.get(alias, 0) + qty_i

    return onway_map, ""
